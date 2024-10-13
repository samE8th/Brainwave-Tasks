'''INVETORY MANAGEMENT SYSTEM'''

import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook, load_workbook
import os

class Product:
    def __init__(self, name, quantity, price):
        self.name = name
        self.quantity = quantity
        self.price = price

class InventoryManager:
    def __init__(self, file_name='inventory.xlsx'):
        self.inventory = {}
        self.file_name = file_name
        self.load_inventory()

    def add_product(self, name, quantity, price):
        if name in self.inventory:
            messagebox.showinfo("Error", "Product already exists.")
        else:
            self.inventory[name] = Product(name, int(quantity), float(price))
            self.save_inventory()
            messagebox.showinfo("Success", f"Product {name} added.")

    def edit_product(self, name, quantity, price):
        if name in self.inventory:
            self.inventory[name].quantity = int(quantity)
            self.inventory[name].price = float(price)
            self.save_inventory()
            messagebox.showinfo("Success", f"Product {name} updated.")
        else:
            messagebox.showinfo("Error", "Product not found.")

    def delete_product(self, name):
        if name in self.inventory:
            del self.inventory[name]
            self.save_inventory()
            messagebox.showinfo("Success", f"Product {name} deleted.")
        else:
            messagebox.showinfo("Error", "Product not found.")

    def low_stock_alert(self, threshold=5):
        low_stock_products = [name for name, product in self.inventory.items() if product.quantity <= threshold]
        if low_stock_products:
            messagebox.showinfo("Low Stock", f"Low stock products: {', '.join(low_stock_products)}")
        else:
            messagebox.showinfo("Low Stock", "All products are sufficiently stocked.")

    def save_inventory(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.append(['Product Name', 'Quantity', 'Price'])

        for product in self.inventory.values():
            ws.append([product.name, product.quantity, product.price])

        wb.save(self.file_name)

    def load_inventory(self):
        if os.path.exists(self.file_name):
            wb = load_workbook(self.file_name)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                name, quantity, price = row
                self.inventory[name] = Product(name, int(quantity), float(price))

class InventoryApp:
    def __init__(self, root):
        self.manager = InventoryManager()
        self.root = root
        self.root.title("Inventory Management System")
        root.configure(bg="ivory3")


        # Product Form
        name_label=tk.Label(root, text="Product Name:",font = ("Helvetica", 12))
        name_label.grid(row=0, column=0, pady=(10, 10))
        name_label.configure(bg="ivory3", fg="black")
        Quantity_label=tk.Label(root, text="Quantity:",font = ("Helvetica", 12))
        Quantity_label.grid(row=1, column=0, pady=(10, 10))
        Quantity_label.configure(bg="ivory3", fg="black")
        price_label=tk.Label(root, text="(₹)Price:",font = ("Helvetica", 12))
        price_label.grid(row=2, column=0, pady=(10, 10))
        price_label.configure(bg="ivory3", fg="black")
        self.name_entry = tk.Entry(root,font=("Arial", 12))
        self.quantity_entry = tk.Entry(root,font=("Arial", 12))
        self.price_entry = tk.Entry(root,font=("Arial", 12))

        self.name_entry.grid(row=0, column=1, pady=(10, 10))
        self.quantity_entry.grid(row=1, column=1, pady=(10, 10))
        self.price_entry.grid(row=2, column=1, pady=(10, 10))

        # Buttons
        Add_btn=tk.Button(root, text="Add Product", command=self.add_product)
        Add_btn.grid(row=3, column=0, pady=10)
        Add_btn.configure(bg="green", fg="white")
        Edit_btn=tk.Button(root, text="Edit Product", command=self.edit_product)
        Edit_btn.grid(row=3, column=1)
        Edit_btn.configure(bg="yellow", fg="black")
        Delete_btn=tk.Button(root, text="Delete Product", command=self.delete_product)
        Delete_btn.grid(row=3, column=2)
        Delete_btn.configure(bg="orange red", fg="black")
        Low_stock_btn=tk.Button(root, text="Low Stock Alert", command=self.low_stock_alert)
        Low_stock_btn.grid(row=4, column=0, pady=10)
        Low_stock_btn.configure(bg="light coral", fg="black")
        Exit_btn=tk.Button(root, text="Exit", command=root.quit)
        Exit_btn.grid(row=4, column=2, pady=10)
        Exit_btn.configure(bg='Red', fg='black')
        Preview_btn=tk.Button(root, text="Preview Inventory", command=self.preview_inventory)
        Preview_btn.grid(row=4, column=1)
        Preview_btn.configure(bg='cyan', fg='black')
        # Treeview to display the product list
        self.tree = ttk.Treeview(root, columns=('Product', 'Quantity', 'Price'), show='headings')
        self.tree.heading('Product', text='Product Name')
        self.tree.heading('Quantity', text='Quantity')
        self.tree.heading('Price', text='Price')
        self.tree.grid(row=5, column=0, columnspan=3, padx=10, pady=10)

    def clear_entries(self):
        self.name_entry.delete(0, tk.END)
        self.quantity_entry.delete(0, tk.END)
        self.price_entry.delete(0, tk.END)

    def add_product(self):
        name = self.name_entry.get()
        quantity = self.quantity_entry.get()
        price = self.price_entry.get()
        if name and quantity and price:
            self.manager.add_product(name, quantity, price)
            self.clear_entries()
            self.preview_inventory()  # Refresh the table
        else:
            messagebox.showinfo("Error", "All fields are required.")

    def edit_product(self):
        name = self.name_entry.get()
        quantity = self.quantity_entry.get()
        price = self.price_entry.get()
        if name and quantity and price:
            self.manager.edit_product(name, quantity, price)
            self.clear_entries()
            self.preview_inventory()  # Refresh the table
        else:
            messagebox.showinfo("Error", "All fields are required.")

    def delete_product(self):
        name = self.name_entry.get()
        if name:
            self.manager.delete_product(name)
            self.clear_entries()
            self.preview_inventory()  # Refresh the table
        else:
            messagebox.showinfo("Error", "Product name is required.")

    def low_stock_alert(self):
        self.manager.low_stock_alert()

    def preview_inventory(self):
        # Clear current treeview contents
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        # Insert updated inventory data into the treeview
        for product in self.manager.inventory.values():
            self.tree.insert('', 'end', values=(product.name, product.quantity, product.price))

# Main Program
if __name__ == "__main__":
    root = tk.Tk()
    app = InventoryApp(root)
    root.mainloop()
